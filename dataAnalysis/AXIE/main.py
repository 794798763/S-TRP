
from openpyxl import load_workbook

import json

def print_hi(name):
    print(f'Hi, {name}')  # Press Ctrl+F8 to toggle the breakpoint.

#根据爬虫excel生成图，将图写入edge.json value.json
def generateGraphFromExcel():
    wb=load_workbook("graphShow.xlsx")
    sheet1=wb[wb.sheetnames[0]]
    accDic = dict()
    accDic_reverse = dict()
    for row in sheet1.iter_rows():
        accFrom=row[0].value
        accTo=row[1].value
        if accFrom is None:
            continue
        if '0x' in accFrom:
            accList=accDic.get(accFrom,[])
            accList.append(accTo)
            accDic[accFrom]=accList

            accList_reverse = accDic_reverse.get(accTo, [])
            accList_reverse.append(accFrom)
            accDic_reverse[accTo] = accList_reverse
    return accDic,accDic_reverse

def writeGraphToJSON(accDic,accDic_reverse):
    edgeJson=json.dumps(accDic)
    edgeJson_reverse=json.dumps(accDic_reverse)
    with open("edge.json",'w') as f1:
        f1.write(edgeJson)
    with open("edge_reverse.json", 'w') as f4:
        f4.write(edgeJson_reverse)

def readGrapyFromJSON():
    with open('edge.json','r') as f1:
        edgesDic=json.load(f1)
    # with open('txHash.json', 'r') as f2:
    #     txHashDic=json.load(f2)
    # with open('blockNum.json', 'r') as f3:
    #     blockNumDic = json.load(f3)
    with open('edge_reverse.json', 'r') as f4:
        edgeDic_reverse = json.load(f4)

    return edgesDic,edgeDic_reverse

def searchPath(targetAddress):
    nodeStake=[]
    pathNodes=set()
    passedNodes=set()
    sourceAddress='0xd4e79226f1e5a7a28abb58f4704e53cd364e8d11'
    nodeStake.append(sourceAddress)
    passedNodes.add(sourceAddress)
    pathNodes.add(sourceAddress)
    pathNodes.add(targetAddress)
    edgesDic,_,_,_=readGrapyFromJSON()
    for i in edgesDic[sourceAddress]:
        nextNode(i,edgesDic,nodeStake,pathNodes,passedNodes)
    return pathNodes

def nextNode(currentNode,edgesDic,nodeStake,pathNodes,passedNodes):
    if currentNode in pathNodes:
        stainPathNodes(nodeStake, pathNodes)
        return
    if currentNode in passedNodes:
        return
    passedNodes.add(currentNode)
    nodeStake.append(currentNode)
    if currentNode in edgesDic.keys():
        for nn in edgesDic[currentNode]:
            nextNode(nn,edgesDic,nodeStake,pathNodes,passedNodes)
    del nodeStake[-1]
    return

def stainPathNodes(nodeStake,pathNode):
    for node in nodeStake:
        pathNode.add(node)

def markExcelTxsInPath(pathNodes,color):
    wb=load_workbook("graphShow.xlsx")
    sheet1=wb[wb.sheetnames[0]]
    for row in sheet1.iter_rows():
        accFrom=row[0].value
        accTo=row[1].value
        if accFrom is None:
            continue
        if '0x' in accFrom:
            if accFrom in pathNodes:
                if accTo in pathNodes:
                    row[2].value=color
    wb.save('graphShow.xlsx')

#广度优先遍历，删掉时序不符的
def washTxsByBlockNumAbandon():
    nodeQueue=[]
    nodeMinBlockNum=dict()
    sourceAddress='0xd4e79226f1e5a7a28abb58f4704e53cd364e8d11'
    nodeQueue.append(sourceAddress)
    nodeMinBlockNum[sourceAddress]='0'
    edgesDic,edgeDic_reverse,txHashDic,edgeBlockNumDic=readGrapyFromJSON()
    #本轮遍历只更新节点最小高度
    while len(nodeQueue)>0 :
        #遍历每一个节点
        cN=nodeQueue.pop(0)
        #该点是不是有出边
        if not cN in edgesDic.keys():
            continue
        cList=edgesDic[cN]
        cNblockNum=nodeMinBlockNum[cN]
        #遍历该点的每一条边
        for i in range(len(cList)):
            edgeBlockNum = edgeBlockNumDic[cN][i]
            if int(edgeBlockNum) < int(cNblockNum):
                continue
            nodeQueue.append(cList[i])
            if cList[i] in nodeMinBlockNum.keys():
                if int(edgeBlockNum)>=int(nodeMinBlockNum[cList[i]]):
                    continue
            nodeMinBlockNum[cList[i]]=edgeBlockNum
    print(nodeMinBlockNum)

def washTxsByTimeStamp():
    #该方法不好用 清洗后只剩下十几笔交易了
    wb=load_workbook("0xd4e79226f1e5a7a28abb58f4704e53cd364e8d11.xlsx")
    sourceAddress="0xd4e79226f1e5a7a28abb58f4704e53cd364e8d11"
    sheet1=wb[wb.sheetnames[0]]
    txList=[]
    #交易全部读入，构成list
    for row in sheet1.iter_rows():
        accFrom = row[2].value
        if accFrom is None:
            continue
        if not '0x' in accFrom:
            continue
        # value=row[4].value
        # if int(value)==0:
        #     continue
        accTo=row[3].value
        txHash=row[0].value
        timeStamp=row[5].value
        txTuple=(txHash,accFrom,accTo,int(timeStamp))
        txList.append(txTuple)
    #排序所有交易
    txList.sort(key= lambda x:x[3])
    listeningNodeSet=set()
    listeningNodeSet.add(sourceAddress)
    removedTxsList=[]
    #顺序遍历交易，开始清洗
    for tx in txList:
        if not tx[1] in listeningNodeSet:
            print("删除交易"+str(tx))
            removedTxsList.append(tx)
    print(len(removedTxsList))
    for i in removedTxsList:
        txList.remove(i)
    washedTxsJSON=json.dumps(txList)
    print(len(txList))
    with open("washedTxs.json",'w') as f1:
        f1.write(washedTxsJSON)

def readWashedTxs():
    with open("washedTxs.json",'r') as f1:
        txList=json.load(f1)
    print(len(txList))

def mergeTxs():
    #合并相同起点、终点的交易
    newEdgeDic={}
    newEdgeDic_reverse={}
    targetNodes=["0x50d1c9771902476076ecfc8b2a83ad6b9355a4c9","0x2faf487a4414fe77e2327f0bf4ae2a264a776ad2","0x871baed4088b863fd6407159f3672d70cd34837d","0xc098b2a3aa256d2140208c3de6543aaef5cd3a94"]
    with open("edge.json",'r') as f1:
        edgeDic=json.load(f1)
    for fromNode in edgeDic.keys():
        for toNode in edgeDic[fromNode]:
            tempList=newEdgeDic.get(fromNode,[])
            if not toNode in tempList:
                tempList.append(toNode)
                newEdgeDic[fromNode]=tempList
                ttlist=newEdgeDic_reverse.get(toNode,[])
                ttlist.append(fromNode)
                newEdgeDic_reverse[toNode]=ttlist

    edgeCount = 0
    nodeSet = set()
    for fromNode in newEdgeDic.keys():
        nodeSet.add(fromNode)
        for toNode in newEdgeDic[fromNode]:
            nodeSet.add(toNode)
            edgeCount = edgeCount + 1
    print("合并重复交易后节点数量" + str(len(nodeSet)))
    print("合并重复交易后无重复边数量" + str(edgeCount))

    newNodeSet=set()
    for n in nodeSet:
        if n in newEdgeDic.keys():
            newNodeSet.add(n)
            continue
        if len(newEdgeDic_reverse.get(n,[]))>1:
            newNodeSet.add(n)
            continue
        if n in targetNodes:
            newNodeSet.add(n)
            continue
        preNlist =newEdgeDic_reverse.get(n,[])
        for preN in preNlist:
            tempList=newEdgeDic.get(preN,[])
            if n in tempList:
                tempList.remove(n)
                newEdgeDic[preN]=tempList
        newEdgeDic_reverse.pop(n)

    edgeCount = 0
    for fromNode in newEdgeDic.keys():
        for toNode in newEdgeDic[fromNode]:
            edgeCount = edgeCount + 1
    print("删除无意义交易和点后节点数量" + str(len(newNodeSet)))
    print("删除无意义交易和点后边数量" + str(edgeCount))
    return newEdgeDic

def markTxsInExcel(edgeDic,color):
    wb=load_workbook("graphShow.xlsx")
    sheet1=wb[wb.sheetnames[0]]
    i=1
    for nodeFrom in edgeDic.keys():
        for nodeTo in edgeDic[nodeFrom]:
        #删除以target为起始点的交易
            if nodeFrom in ["0x50d1c9771902476076ecfc8b2a83ad6b9355a4c9","0x2faf487a4414fe77e2327f0bf4ae2a264a776ad2","0x871baed4088b863fd6407159f3672d70cd34837d","0xc098b2a3aa256d2140208c3de6543aaef5cd3a94"]:
                continue
            sheet1.cell(i,1,nodeFrom)
            sheet1.cell(i,2,nodeTo)
            sheet1.cell(i,3,color)
            i=i+1
    wb.save('graphShow.xlsx')

def searchPathByDic(edgesDic,targetAddress):
    nodeStake=[]
    pathNodes=set()
    passedNodes=set()
    sourceAddress='0x098b716b8aaf21512996dc57eb0615e2383e2f96'
    nodeStake.append(sourceAddress)
    passedNodes.add(sourceAddress)
    pathNodes.add(sourceAddress)
    pathNodes.add(targetAddress)
    for i in edgesDic[sourceAddress]:
        nextNode(i,edgesDic,nodeStake,pathNodes,passedNodes)
    print(len(pathNodes))
    return pathNodes

def replenishEdges(number):
    wb=load_workbook("0xd4e79226f1e5a7a28abb58f4704e53cd364e8d11.xlsx")
    sheet1=wb[wb.sheetnames[0]]

    wb2 = load_workbook("greyGraphShow.xlsx")
    sheet2 = wb2[wb2.sheetnames[0]]

    color = "grey"
    count=0
    for row in sheet1.iter_rows():
        if count>number:
            break
        count=count+1

        accFrom = row[2].value
        if accFrom is None:
            continue
        if not '0x' in accFrom:
            continue
        accTo=row[3].value

        sheet2.append([accFrom,accTo,color])
    wb2.save('greyGraphShow.xlsx')

def findCore(nodeDic,nodeDic_reverse):
    coreSet=set()
    for nodeFrom in nodeDic.keys():
        if len(nodeDic[nodeFrom])+len(nodeDic_reverse[nodeFrom])>20:
            coreSet.add(nodeFrom)
        for nodeTo in nodeDic[nodeFrom]:
            if len(nodeDic[nodeFrom])+len(nodeDic_reverse[nodeTo])>20:
                coreSet.add(nodeTo)
    return coreSet

if __name__ == '__main__':
    # re,_=generateGraphFromExcel()
    # re2=mergeTxs()
    # markTxsInExcel(re2,"grey")

    # writeGraphToJSON(re,re2)





    re,_=generateGraphFromExcel()
    # re=readGrapyFromJSON()
    re2=searchPathByDic(re,"0xc098b2a3aa256d2140208c3de6543aaef5cd3a94")
    markExcelTxsInPath(re2,"red")

    # replenishEdges(2000)


    # washTxsByTimeStamp()
    # readWashedTxs()
    # mergeTxs()
    # generateGraphFromExcel()
    # readGrapyFromJSON()
    # print(searchPath())
    # washTxsByBlockNum()
    # markExcelTxsInPath()